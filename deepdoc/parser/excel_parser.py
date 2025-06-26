#  Licensed under the Apache License, Version 2.0 (the "License");
#  you may not use this file except in compliance with the License.
#  You may obtain a copy of the License at
#
#      http://www.apache.org/licenses/LICENSE-2.0
#
#  Unless required by applicable law or agreed to in writing, software
#  distributed under the License is distributed on an "AS IS" BASIS,
#  WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#  See the License for the specific language governing permissions and
#  limitations under the License.
#

import logging
import sys
from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook

from rag.nlp import find_codec


class RAGFlowExcelParser:

    @staticmethod
    def _load_excel_to_workbook(file_like_object):
        if isinstance(file_like_object, bytes):
            file_like_object = BytesIO(file_like_object)

        # Read first 4 bytes to determine file type
        file_like_object.seek(0)
        file_head = file_like_object.read(4)
        file_like_object.seek(0)

        if not (file_head.startswith(b'PK\x03\x04') or file_head.startswith(b'\xD0\xCF\x11\xE0')):
            logging.info("****wxy: Not an Excel file, converting CSV to Excel Workbook")

            try:
                file_like_object.seek(0)
                df = pd.read_csv(file_like_object)
                return RAGFlowExcelParser._dataframe_to_workbook(df)

            except Exception as e_csv:
                raise Exception(f"****wxy: Failed to parse CSV and convert to Excel Workbook: {e_csv}")

        try:
            return load_workbook(file_like_object,data_only= True)  # data_only, don't include formula
        except Exception as e:
            logging.info(f"****wxy: openpyxl load error: {e}, try pandas instead")
            try:
                file_like_object.seek(0)
                df = pd.read_excel(file_like_object)
                return RAGFlowExcelParser._dataframe_to_workbook(df)
            except Exception as e_pandas:
                raise Exception(f"****wxy: pandas.read_excel error: {e_pandas}, original openpyxl error: {e}")

    @staticmethod
    def _dataframe_to_workbook(df):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        for col_num, column_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_num, value=column_name)

        for row_num, row in enumerate(df.values, 2):
            for col_num, value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        return wb

    def html(self, fnm, chunk_rows=256):
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)
        tb_chunks = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue

            tb_rows_0 = "<tr>"
            for t in list(rows[0]):
                tb_rows_0 += f"<th>{t.value}</th>"
            tb_rows_0 += "</tr>"

            for chunk_i in range((len(rows) - 1) // chunk_rows + 1):
                tb = ""
                tb += f"<table><caption>{sheetname}</caption>"
                tb += tb_rows_0
                for r in list(
                    rows[1 + chunk_i * chunk_rows: 1 + (chunk_i + 1) * chunk_rows]
                ):
                    tb += "<tr>"
                    for i, c in enumerate(r):
                        if c.value is None:
                            tb += "<td></td>"
                        else:
                            tb += f"<td>{c.value}</td>"
                    tb += "</tr>"
                tb += "</table>\n"
                tb_chunks.append(tb)

        return tb_chunks

    def __call__(self, fnm):
        file_like_object = BytesIO(fnm) if not isinstance(fnm, str) else fnm
        wb = RAGFlowExcelParser._load_excel_to_workbook(file_like_object)

        res = []
        for sheetname in wb.sheetnames:
            ws = wb[sheetname]
            rows = list(ws.rows)
            if not rows:
                continue
            ti = list(rows[0])
            for r in list(rows[1:]):
                fields = []
                for i, c in enumerate(r):
                    if not c.value:
                        continue
                    t = str(ti[i].value) if i < len(ti) else ""
                    t += ("：" if t else "") + str(c.value)
                    fields.append(t)
                line = "; ".join(fields)
                if sheetname.lower().find("sheet") < 0:
                    line += " ——" + sheetname
                res.append(line)
        return res

    @staticmethod
    def row_number(fnm, binary):
        if fnm.split(".")[-1].lower().find("xls") >= 0:
            wb = RAGFlowExcelParser._load_excel_to_workbook(BytesIO(binary))
            total = 0
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                total += len(list(ws.rows))
            return total

        if fnm.split(".")[-1].lower() in ["csv", "txt"]:
            encoding = find_codec(binary)
            txt = binary.decode(encoding, errors="ignore")
            return len(txt.split("\n"))


if __name__ == "__main__":
    import argparse
    import os
    import time
    from pprint import pprint
    
    # 配置日志
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    logger = logging.getLogger("ExcelParserTest")
    
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description='Test RAGFlowExcelParser functionality')
    parser.add_argument('--file_path',
                        default='/Users/wangshuang/Downloads/9月请假记录.xlsx',
                        help='Path to Excel, CSV or TXT file to parse',
                        )
    parser.add_argument('--html', action='store_true', help='Convert to HTML chunks')
    parser.add_argument('--chunk-rows', type=int, default=256, help='Number of rows per HTML chunk (default: 256)')
    parser.add_argument('--text', action='store_true', help='Extract text content')
    parser.add_argument('--all', default=True, action='store_true', help='Run all tests')
    args = parser.parse_args()
    
    # 验证文件是否存在
    if not os.path.exists(args.file_path):
        logger.error(f"File not found: {args.file_path}")
        sys.exit(1)
    
    # 创建解析器实例
    parser = RAGFlowExcelParser()
    logger.info(f"Testing file: {args.file_path}")
    
    # 读取文件内容
    with open(args.file_path, 'rb') as f:
        file_content = f.read()
    
    # 测试1: 加载为 workbook
    try:
        logger.info("TEST 1: Loading file as workbook")
        start_time = time.time()
        wb = RAGFlowExcelParser._load_excel_to_workbook(BytesIO(file_content))
        elapsed = time.time() - start_time
        
        sheet_info = []
        total_rows = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.rows)
            row_count = len(rows)
            total_rows += row_count
            col_count = len(rows[0]) if row_count > 0 else 0
            sheet_info.append({
                "sheet_name": sheet_name,
                "rows": row_count,
                "columns": col_count
            })
        
        logger.info(f"Workbook loaded successfully in {elapsed:.2f} seconds")
        logger.info(f"Sheets: {len(wb.sheetnames)}")
        logger.info(f"Total rows: {total_rows}")
        logger.info("Sheet details:")
        for info in sheet_info:
            logger.info(f"  - {info['sheet_name']}: {info['rows']} rows, {info['columns']} columns")
            
        # 显示第一个表的前5行样本数据(如果有)
        if wb.sheetnames and sheet_info[0]["rows"] > 0:
            logger.info(f"Sample data from first sheet ({wb.sheetnames[0]}):")
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.rows)
            for i, row in enumerate(rows[:5]):
                row_data = [cell.value for cell in row]
                logger.info(f"  Row {i+1}: {row_data}")
        
    except Exception as e:
        logger.error(f"Failed to load workbook: {e}")
        import traceback
        traceback.print_exc()
    
    # 测试2: 转换为 HTML 块
    if args.html or args.all:
        try:
            logger.info(f"\nTEST 2: Converting to HTML chunks (chunk_rows={args.chunk_rows})")
            start_time = time.time()
            html_chunks = parser.html(file_content, chunk_rows=args.chunk_rows)
            elapsed = time.time() - start_time
            
            logger.info(f"HTML conversion completed in {elapsed:.2f} seconds")
            logger.info(f"Generated {len(html_chunks)} HTML chunks")
            
            # 显示第一个HTML块的样本
            if html_chunks:
                sample_length = min(200, len(html_chunks[0]))
                logger.info(f"Sample of first HTML chunk (first {sample_length} chars):")
                logger.info(html_chunks[0][:sample_length] + "...")
                
                # 保存HTML样本到文件以便查看
                sample_file = f"excel_parser_sample_{int(time.time())}.html"
                with open(sample_file, 'w', encoding='utf-8') as f:
                    f.write("<html><body>\n")
                    for i, chunk in enumerate(html_chunks[:3]):  # 保存前3个块
                        f.write(f"<h2>Chunk {i+1}</h2>\n")
                        f.write(chunk)
                        f.write("<hr>\n")
                    f.write("</body></html>")
                logger.info(f"Saved sample HTML to {sample_file}")
                
        except Exception as e:
            logger.error(f"Failed to convert to HTML: {e}")
            import traceback
            traceback.print_exc()
    
    # 测试3: 提取文本内容
    if args.text or args.all:
        try:
            logger.info("\nTEST 3: Extracting text content")
            start_time = time.time()
            text_lines = parser(file_content)
            elapsed = time.time() - start_time
            
            logger.info(f"Text extraction completed in {elapsed:.2f} seconds")
            logger.info(f"Extracted {len(text_lines)} text lines")
            
            # 显示前10行文本样本
            if text_lines:
                logger.info("Sample text lines (first 10):")
                for i, line in enumerate(text_lines[:10]):
                    logger.info(f"  Line {i+1}: {line[:100]}..." if len(line) > 100 else f"  Line {i+1}: {line}")
                
                # 保存文本样本到文件
                sample_file = f"excel_parser_text_{int(time.time())}.txt"
                with open(sample_file, 'w', encoding='utf-8') as f:
                    for line in text_lines:
                        f.write(line + "\n")
                logger.info(f"Saved extracted text to {sample_file}")
                
        except Exception as e:
            logger.error(f"Failed to extract text: {e}")
            import traceback
            traceback.print_exc()
    
    # 测试4: 行数计算
    try:
        logger.info("\nTEST 4: Counting rows")
        start_time = time.time()
        row_count = RAGFlowExcelParser.row_number(args.file_path, file_content)
        elapsed = time.time() - start_time
        
        logger.info(f"Row counting completed in {elapsed:.2f} seconds")
        logger.info(f"Total rows: {row_count}")
        
    except Exception as e:
        logger.error(f"Failed to count rows: {e}")
        import traceback
        traceback.print_exc()
    
    logger.info("\nAll tests completed")
